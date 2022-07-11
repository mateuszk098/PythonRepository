from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

# wb = Workbook()
# ws = wb.active
# ws.title = "Data"

# # Dodanie wiersza z nazwami kolumn
# ws.append(["Pierwszy", "Drugi", "Trzeci"])

# wb.save("Pierwszy.xlsx")

wb = load_workbook("Pierwszy.xlsx")
ws = wb.active

# Wypełnianie komórek wartościami, zczytywanie wartości
for row in range(1, 11):
    for col in range(1, 5):
        char = get_column_letter(col)
        #print(ws[char + str(row)].value)
        ws[char + str(row)] = char + str(row)

wb.save("Pierwszy.xlsx")
