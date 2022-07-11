from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

wb = load_workbook("Pierwszy.xlsx")
ws = wb.active

# Łączenie komórek
ws.merge_cells("A1:D1")
ws.unmerge_cells("A1:D1")

# Wstawianie wierszy
ws.insert_rows(7)
# Wstawianie komulmn
ws.insert_cols(2)

# Usuwanie wierszy
ws.delete_rows(1)
# Wstawianie komulmn
ws.delete_cols(3)

# Przeniesienie obszaru komórek o 5 wierszy w prawo i 5 kolumn w dół, analogicznie z minusem oznaczałoby to przesienie w lewo i w górę
ws.move_range("A1:D3", rows=5, cols=5)

wb.save("Pierwszy.xlsx")
