# Dokumentacja biblioteki https://openpyxl.readthedocs.io/en/stable/

from openpyxl import Workbook, load_workbook

# Otwieramy plik excel
wb = load_workbook("test.xlsx")

# Wybieramy aktywny arkusz
ws = wb.active

# Wypisanie wartości danej komórki na ekran
# print(ws["A1"].value)

# Zmiana wartości danej komórki
# ws["A2"].value = "Jim"

# Zapisanie arkusza
# wb.save("test.xlsx")

# Wypisanie nazw arkuszy
# print(wb.sheetnames)

# Wybranie 2 arkusza jako aktywny
# ws = wb["Sheet2"]

# Tworzenie nowego arkusza
# wb.create_sheet("Test")
# print(wb.sheetnames)
