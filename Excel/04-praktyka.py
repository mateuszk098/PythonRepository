# Wczytamy oceny studentów z pliku json, wpisujemy je do excela i liczymy średnią a także kolorujemy nagłówki

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
import json

# Czytamy dane z jsona
with open("data.json") as json_file:
    data = json.load(json_file)

# Excel
wb = Workbook()
ws = wb.active
ws.title = "Students Grades"

# Lista nagłówków
headings = ["Name"] + list(data["Joe"].keys())
# Dodajemy nagłówki do excela
ws.append(headings)

# Dla każdej osoby w data dodajemy jej oceny do excela
for person in data:
    grades = list(data[person].values())
    ws.append([person] + grades)

# Liczymy średnią ocen studentów
# Pętla po kolumnach, zaczynamy od B czyli 2 i robimy tak po wszystkich kolmnach mających oceny
for col in range(2, len(data["Joe"]) + 2):
    char = get_column_letter(col)
    # Wstawiamy średnie pod ostatnią osobą (+2 bo dodaliśmy jeszcze nagłówki) char + '2' to B2, char + str(len(data) + 1) to w tym przypadku B6
    ws[char + str(len(data) + 2)
       ] = f"=SUM({char + '2'}:{char + str(len(data) + 1)})/{len(data)}"

# Kolorujemy nagłówki
for col in range(1, len(data["Joe"]) + 2):
    char = get_column_letter(col)
    ws[char + '1'].font = Font(bold=True, color="00FF0000")


# Zapisujemy
wb.save("From_Python.xlsx")
